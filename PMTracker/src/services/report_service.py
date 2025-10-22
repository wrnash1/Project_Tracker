from typing import List, Dict, Optional
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
from pathlib import Path
import tempfile

class ReportService:
    """Advanced report generation service"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._create_custom_styles()

    def _create_custom_styles(self):
        """Create custom report styles"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#DC2626'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))

        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#6B7280'),
            spaceAfter=12,
            alignment=TA_CENTER
        ))

    def generate_executive_summary(self, projects: List[Dict], output_path: str = None) -> str:
        """Generate executive summary report with charts"""
        if not output_path:
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            output_path = temp_file.name
            temp_file.close()

        doc = SimpleDocTemplate(output_path, pagesize=letter)
        elements = []

        # Title
        title = Paragraph("Executive Project Summary", self.styles['CustomTitle'])
        elements.append(title)

        # Date
        date_text = f"Report Generated: {datetime.now().strftime('%B %d, %Y')}"
        elements.append(Paragraph(date_text, self.styles['CustomSubtitle']))
        elements.append(Spacer(1, 0.3 * inch))

        # Summary statistics
        total_projects = len(projects)
        active_projects = len([p for p in projects if p.get('PROJECT_STATUS') == 'ACTIVE'])
        completed_projects = len([p for p in projects if p.get('PROJECT_STATUS') == 'COMPLETED'])
        total_budget = sum([p.get('BUDGET', 0) for p in projects])
        total_actual = sum([p.get('ACTUAL_COST', 0) for p in projects])

        stats_data = [
            ['Metric', 'Value'],
            ['Total Projects', str(total_projects)],
            ['Active Projects', str(active_projects)],
            ['Completed Projects', str(completed_projects)],
            ['Total Budget', f'${total_budget:,.2f}'],
            ['Total Actual Cost', f'${total_actual:,.2f}'],
            ['Budget Variance', f'${total_budget - total_actual:,.2f}']
        ]

        stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))

        elements.append(stats_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Project status distribution
        elements.append(Paragraph("Project Status Distribution", self.styles['Heading2']))
        elements.append(Spacer(1, 0.2 * inch))

        # Add pie chart
        pie_chart = self._create_status_pie_chart(projects)
        elements.append(pie_chart)
        elements.append(PageBreak())

        # Top projects by budget
        elements.append(Paragraph("Top 10 Projects by Budget", self.styles['Heading2']))
        elements.append(Spacer(1, 0.2 * inch))

        sorted_projects = sorted(projects, key=lambda p: p.get('BUDGET', 0), reverse=True)[:10]
        project_data = [['Project Number', 'Project Name', 'Budget', 'Status']]

        for project in sorted_projects:
            project_data.append([
                project.get('PROJECT_NUMBER', 'N/A'),
                project.get('PROJECT_NAME', 'N/A')[:40],
                f"${project.get('BUDGET', 0):,.2f}",
                project.get('PROJECT_STATUS', 'N/A')
            ])

        project_table = Table(project_data, colWidths=[1.2*inch, 2.5*inch, 1.5*inch, 1.3*inch])
        project_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(project_table)

        # Build PDF
        doc.build(elements)
        return output_path

    def _create_status_pie_chart(self, projects: List[Dict]) -> Drawing:
        """Create pie chart for project status"""
        status_counts = {}
        for project in projects:
            status = project.get('PROJECT_STATUS', 'Unknown')
            status_counts[status] = status_counts.get(status, 0) + 1

        drawing = Drawing(400, 200)
        pie = Pie()
        pie.x = 150
        pie.y = 50
        pie.width = 100
        pie.height = 100

        pie.data = list(status_counts.values())
        pie.labels = list(status_counts.keys())
        pie.slices.strokeWidth = 0.5

        colors_list = [
            colors.HexColor('#DC2626'),  # Red
            colors.HexColor('#3B82F6'),  # Blue
            colors.HexColor('#10B981'),  # Green
            colors.HexColor('#F59E0B'),  # Amber
            colors.HexColor('#8B5CF6'),  # Purple
        ]

        for i, color in enumerate(colors_list[:len(pie.data)]):
            pie.slices[i].fillColor = color

        drawing.add(pie)
        return drawing

    def generate_budget_analysis(self, projects: List[Dict], output_path: str = None) -> str:
        """Generate detailed budget analysis report"""
        if not output_path:
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            output_path = temp_file.name
            temp_file.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget Analysis"

        # Header styling
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Headers
        headers = ['Project Number', 'Project Name', 'Budget', 'Actual Cost',
                  'Variance', 'Variance %', 'Status', 'PM Name']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        row = 2
        for project in projects:
            budget = project.get('BUDGET', 0)
            actual = project.get('ACTUAL_COST', 0)
            variance = budget - actual
            variance_pct = (variance / budget * 100) if budget > 0 else 0

            ws.cell(row=row, column=1, value=project.get('PROJECT_NUMBER', 'N/A'))
            ws.cell(row=row, column=2, value=project.get('PROJECT_NAME', 'N/A'))
            ws.cell(row=row, column=3, value=budget)
            ws.cell(row=row, column=4, value=actual)
            ws.cell(row=row, column=5, value=variance)
            ws.cell(row=row, column=6, value=variance_pct)
            ws.cell(row=row, column=7, value=project.get('PROJECT_STATUS', 'N/A'))
            ws.cell(row=row, column=8, value=project.get('PM_NAME', 'N/A'))

            # Format currency columns
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=6).number_format = '0.00%'

            # Color code variance
            variance_cell = ws.cell(row=row, column=5)
            if variance < 0:
                variance_cell.font = Font(color="DC2626")  # Red for over budget
            else:
                variance_cell.font = Font(color="10B981")  # Green for under budget

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20

        # Add chart
        chart = BarChart()
        chart.title = "Budget vs Actual Cost - Top 10 Projects"
        chart.y_axis.title = 'Amount ($)'
        chart.x_axis.title = 'Project'

        data = Reference(ws, min_col=3, min_row=1, max_row=min(11, row), max_col=4)
        cats = Reference(ws, min_col=1, min_row=2, max_row=min(11, row))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20

        ws.add_chart(chart, "J2")

        wb.save(output_path)
        return output_path

    def generate_timeline_report(self, projects: List[Dict], output_path: str = None) -> str:
        """Generate project timeline report"""
        if not output_path:
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            output_path = temp_file.name
            temp_file.close()

        doc = SimpleDocTemplate(output_path, pagesize=A4)
        elements = []

        # Title
        title = Paragraph("Project Timeline Report", self.styles['CustomTitle'])
        elements.append(title)
        elements.append(Spacer(1, 0.3 * inch))

        # Sort projects by start date
        sorted_projects = sorted(
            [p for p in projects if p.get('START_DATE')],
            key=lambda p: p.get('START_DATE')
        )

        # Timeline data
        timeline_data = [['Project', 'Start Date', 'End Date', 'Duration (Days)', 'Status']]

        for project in sorted_projects[:20]:  # Top 20 projects
            start_date = project.get('START_DATE')
            end_date = project.get('END_DATE')

            if start_date and end_date:
                duration = (end_date - start_date).days
            else:
                duration = 'N/A'

            timeline_data.append([
                project.get('PROJECT_NUMBER', 'N/A'),
                start_date.strftime('%Y-%m-%d') if start_date else 'N/A',
                end_date.strftime('%Y-%m-%d') if end_date else 'N/A',
                str(duration) if isinstance(duration, int) else duration,
                project.get('PROJECT_STATUS', 'N/A')
            ])

        timeline_table = Table(timeline_data, colWidths=[1.5*inch, 1.3*inch, 1.3*inch, 1.2*inch, 1.2*inch])
        timeline_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(timeline_table)
        doc.build(elements)
        return output_path

    def generate_custom_report(self, title: str, data: List[Dict],
                              columns: List[str], output_format: str = 'pdf') -> str:
        """Generate custom report with specified columns"""
        if output_format == 'pdf':
            return self._generate_custom_pdf(title, data, columns)
        else:
            return self._generate_custom_excel(title, data, columns)

    def _generate_custom_pdf(self, title: str, data: List[Dict], columns: List[str]) -> str:
        """Generate custom PDF report"""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        output_path = temp_file.name
        temp_file.close()

        doc = SimpleDocTemplate(output_path, pagesize=letter)
        elements = []

        # Title
        elements.append(Paragraph(title, self.styles['CustomTitle']))
        elements.append(Spacer(1, 0.3 * inch))

        # Create table data
        table_data = [columns]
        for item in data:
            row = [str(item.get(col, 'N/A')) for col in columns]
            table_data.append(row)

        # Calculate column widths
        col_width = 6.5 * inch / len(columns)
        col_widths = [col_width] * len(columns)

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table)
        doc.build(elements)
        return output_path

    def _generate_custom_excel(self, title: str, data: List[Dict], columns: List[str]) -> str:
        """Generate custom Excel report"""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        output_path = temp_file.name
        temp_file.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limit

        # Headers
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        for row, item in enumerate(data, 2):
            for col, column in enumerate(columns, 1):
                ws.cell(row=row, column=col, value=item.get(column, 'N/A'))

        wb.save(output_path)
        return output_path
